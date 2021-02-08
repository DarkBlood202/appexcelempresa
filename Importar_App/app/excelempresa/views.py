from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import redirect, reverse

from .forms import UploadFileForm, FechaForm
from .functions import handle_uploaded_file, remove_uploaded_file_on, cells_to_dict
from .functions import insertar_repartidores, insertar_restaurantes, insertar_pedidos, insertar_envios

from django.template.loader import get_template
from .functions import render_to_pdf

import openpyxl
import datetime

from .models import Restaurante, Repartidor, Pedido, Envio
from .models import Constants

# Create your views here.
# def index(request):
#    # return HttpResponse("<h1>Index</h1><hr>")


def upload(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            storage_path = 'excelempresa/upload/'
            handle_uploaded_file(request.FILES['file'],storage_path)
            
            # Extraer data desde Excel
            data = openpyxl.load_workbook(storage_path + str(request.FILES['file']), data_only=True)
            
            t_pedidos = data["PEDIDOS"]
            t_envios = data["ENVIOS"]
            t_restaurantes = data["RESTAURANTES"]
            t_repartidores = data["REPARTIDORES"]

            f_pedidos = t_pedidos.max_row - 1
            f_envios = t_envios.max_row - 1
            f_restaurantes = t_restaurantes.max_row - 1
            f_repartidores = t_repartidores.max_row - 1

            c_pedidos = t_pedidos.max_column
            c_envios = t_envios.max_column
            c_restaurantes = t_restaurantes.max_column
            c_repartidores = t_repartidores.max_column

            dict_pedidos = cells_to_dict(t_pedidos,f_pedidos,c_pedidos)
            dict_envios = cells_to_dict(t_envios,f_envios,c_pedidos)
            dict_restaurantes = cells_to_dict(t_restaurantes,f_restaurantes,c_restaurantes)
            dict_repartidores = cells_to_dict(t_repartidores,f_repartidores,c_repartidores)

            # Crear objetos en base a los datos extraidos y hoja de datos
            INDICES_PEDIDO = [0,1,2,3,5,6,8,12,20]
            LLAVES_PEDIDO = {0:'restaurante',1:'numero',2:'task_id',3:'repartidor',5:'estado',
                6:'subtotal_pedido',8:'tipo',12:'total_pedido',20:'fecha'}
            

            INDICES_ENVIO = [0,2,3,6,8]
            LLAVES_ENVIO = {0:'pedido',2:'repartidor',3:'coste_envio',6:'fecha',
                8:'distancia_metros'}


            INDICES_RESTAURANTE = [0,1,2,3,4,7,8,9,11,12,13,14,15,16,17,18,19,20]
            LLAVES_RESTAURANTE = {0:'id',1:'nombre',2:'cantidad_pedidos',3:'cantidad_recaudada',
                4:'precio_base_envio',7:'gastos_envio',8:'cantidad_transferir',9:'razon_social',
                11:'nif_cif',12:'direccion',13:'telefono',14:'email',15:'poblacion',16:'provincia',
                17:'codigo_postal',18:'cuenta_bancaria',19:'importe_sin_iva',20:'importe_iva'}

            INDICES_REPARTIDOR = [0,1,2,7,8,9,10,11,12,13,14,15,16,17]
            LLAVES_REPARTIDOR = {0:'id',1:'nombre',2:'cantidad_repartos',7:'cantidad_transferir',
                8:'telefono',9:'nif_cif',10:'direccion',11:'email',12:'poblacion',13:'provincia',
                14:'codigo_postal',15:'cuenta_bancaria',16:'importe_iva',17:'importe_sin_iva'}

            try:
                insertar_repartidores(f_repartidores,INDICES_REPARTIDOR,LLAVES_REPARTIDOR,dict_repartidores)
                insertar_restaurantes(f_restaurantes,INDICES_RESTAURANTE,LLAVES_RESTAURANTE,dict_restaurantes)
                insertar_pedidos(f_pedidos,INDICES_PEDIDO,LLAVES_PEDIDO,dict_pedidos)
                insertar_envios(f_envios,INDICES_ENVIO,LLAVES_ENVIO,dict_envios)
            except Exception as ex:
                return HttpResponse("<h1>Error: Couldn't load data</h1><p>{}</p>".format(ex))
            finally:
                remove_uploaded_file_on(storage_path)

            #return redirect(reverse('admin.site.urls'))
            return HttpResponse("Success.")
    else:
        form = UploadFileForm()
        return render(request, 'excelempresa/upload.html', {'form':form})


def liquidacion_rp(request, id_repartidor):
    template = get_template('excelempresa/liquidacion_rp.html')
    q = get_object_or_404(Repartidor, id=id_repartidor)

    pp = Pedido.objects.filter(repartidor=Repartidor.objects.get(id=id_repartidor))
    ee = Envio.objects.filter(repartidor=Repartidor.objects.get(id=id_repartidor))
    
    try:
        fecha_inicio = Constants.objects.get(id=1).fecha_inicio
        fecha_fin = Constants.objects.get(id=1).fecha_fin
    except:
        c = Constants()
        c.save()

    importe_recaudado = 0
    cantidad_repartos = 0
    for e in ee:
        if e.pedido.fecha >= fecha_inicio and e.pedido.fecha <= fecha_fin:
            importe_recaudado += e.coste_envio
            cantidad_repartos += 1

    importe_iva = round(float(importe_recaudado) * 0.21,2)
    importe_sin_iva = round(float(importe_recaudado) - importe_iva,2)

    context = {'repartidor': q, 'lista_pedidos': pp, 'lista_envios':ee,
        'importe_recaudado': importe_recaudado, 'importe_iva': importe_iva,
        'importe_sin_iva': importe_sin_iva, 'cantidad_repartos': cantidad_repartos,
        'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
    html = template.render(context)
    pdf = render_to_pdf('excelempresa/liquidacion_rp.html', context)

    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "Hoja de liquidacion de repartidor - {}.pdf".format(q.nombre)
        content = "inline; filename={}".format(filename)

        download = request.GET.get("download")
        if download:
            content = "attachment; filename={}".format(filename)
        
        response['Content-Disposition'] = content

        return response
    return HttpResponseNotFound("<h1>Repartidor solicitado no existe.</h1>")


def liquidacion_rs(request, id_restaurante):
    template = get_template('excelempresa/liquidacion_rs.html')
    q = get_object_or_404(Restaurante, id=id_restaurante)

    pp = Pedido.objects.filter(restaurante=Restaurante.objects.get(id=id_restaurante))
    ee = Envio.objects.all()

    er = []
    for e in ee:
        if e.pedido in pp:
            er.append(e)

    try:
        fecha_inicio = Constants.objects.get(id=1).fecha_inicio
        fecha_fin = Constants.objects.get(id=1).fecha_fin
    except:
        c = Constants()
        c.save()

    total_recaudado = 0
    cantidad_pedidos = 0
    gastos_envio = 0 # suma de costes restaurante por pedido/envio
    for p in pp:
        if p.fecha >= fecha_inicio and p.fecha <= fecha_fin:
            total_recaudado += p.subtotal_pedido
            cantidad_pedidos += 1

    for envio in er:
        if envio.pedido.fecha >= fecha_inicio and envio.pedido.fecha <= fecha_fin:
            gastos_envio += (envio.coste_envio - envio.pedido.restaurante.precio_base_envio)

    importe_final = total_recaudado - gastos_envio
    importe_sin_iva = round(float(importe_final) / 1.1,2)
    importe_iva = round(float(importe_final) - importe_sin_iva,2)

    try:
        porcentaje_gastos = "{:.2%}".format(gastos_envio / total_recaudado)
    except:
        porcentaje_gastos = "0%"

    context = {'restaurante': q, 'lista_pedidos': pp, 'lista_envios': er,
        'total_recaudado': total_recaudado, 'cantidad_pedidos': cantidad_pedidos,
        'gastos_envio': gastos_envio, 'importe_final': importe_final,
        'importe_sin_iva': importe_sin_iva, 'importe_iva': importe_iva,
        'porcentaje_gastos': porcentaje_gastos,
        'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
    
    html = template.render(context)
    pdf = render_to_pdf('excelempresa/liquidacion_rs.html', context)

    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "Hoja de liquidacion de restaurante - {}.pdf".format(q.nombre)
        content = "inline; filename={}".format(filename)

        download = request.GET.get("download")
        if download:
            content = "attachment; filename={}".format(filename)
        
        response['Content-Disposition'] = content

        return response
    return HttpResponseNotFound("<h1>Restaurante solicitado no existe.</h1>")


# Vistas admin
def vista_importar_repartidores(request):
    if request.method  == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            storage_path = 'excelempresa/upload/'
            handle_uploaded_file(request.FILES['file'],storage_path)

            data = openpyxl.load_workbook(storage_path + str(request.FILES['file']), data_only=True)
            t_repartidores = data["REPARTIDORES"]

            f_repartidores = t_repartidores.max_row - 1
            c_repartidores = t_repartidores.max_column

            dict_repartidores = cells_to_dict(t_repartidores,f_repartidores,c_repartidores)

            INDICES_REPARTIDOR = [0,1,2,7,8,9,10,11,12,13,14,15,16,17]
            LLAVES_REPARTIDOR = {0:'id',1:'nombre',2:'cantidad_repartos',7:'cantidad_transferir',
                8:'telefono',9:'nif_cif',10:'direccion',11:'email',12:'poblacion',13:'provincia',
                14:'codigo_postal',15:'cuenta_bancaria',16:'importe_iva',17:'importe_sin_iva'}

            try:
                insertar_repartidores(f_repartidores,INDICES_REPARTIDOR,LLAVES_REPARTIDOR,dict_repartidores)
                remove_uploaded_file_on(storage_path)
                return redirect('../')
            except Exception as ex:
                return HttpResponse("<h1>Error: Couldn't load data</h1><p>{}</p>".format(ex))
            finally:
                remove_uploaded_file_on(storage_path)
    else:
        form = UploadFileForm()
        return render(request, 'excelempresa/import_intermediate.html', {'form': form, 'modelo':'repartidores'})

def vista_importar_restaurantes(request):
    if request.method  == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            storage_path = 'excelempresa/upload/'
            handle_uploaded_file(request.FILES['file'],storage_path)

            data = openpyxl.load_workbook(storage_path + str(request.FILES['file']), data_only=True)
            t_restaurantes = data["RESTAURANTES"]

            f_restaurantes = t_restaurantes.max_row - 1
            c_restaurantes = t_restaurantes.max_column

            dict_restaurantes = cells_to_dict(t_restaurantes,f_restaurantes,c_restaurantes)

            INDICES_RESTAURANTE = [0,1,2,3,4,7,8,9,11,12,13,14,15,16,17,18,19,20]
            LLAVES_RESTAURANTE = {0:'id',1:'nombre',2:'cantidad_pedidos',3:'cantidad_recaudada',
                4:'precio_base_envio',7:'gastos_envio',8:'cantidad_transferir',9:'razon_social',
                11:'nif_cif',12:'direccion',13:'telefono',14:'email',15:'poblacion',16:'provincia',
                17:'codigo_postal',18:'cuenta_bancaria',19:'importe_sin_iva',20:'importe_iva'}     

            try:
                insertar_restaurantes(f_restaurantes,INDICES_RESTAURANTE,LLAVES_RESTAURANTE,dict_restaurantes)
                remove_uploaded_file_on(storage_path)
                return redirect('../')
            except Exception as ex:
                return HttpResponse("<h1>Error: Couldn't load data</h1><p>{}</p>".format(ex))
            finally:
                remove_uploaded_file_on(storage_path)
    else:
        form = UploadFileForm()
        return render(request, 'excelempresa/import_intermediate.html', {'form': form, 'modelo':'restaurantes'})

def vista_importar_pedidos(request):
    if request.method  == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            storage_path = 'excelempresa/upload/'
            handle_uploaded_file(request.FILES['file'],storage_path)

            data = openpyxl.load_workbook(storage_path + str(request.FILES['file']), data_only=True)
            t_pedidos = data["PEDIDOS"]

            f_pedidos = t_pedidos.max_row - 1
            c_pedidos = t_pedidos.max_column

            dict_pedidos = cells_to_dict(t_pedidos,f_pedidos,c_pedidos)

            INDICES_PEDIDO = [0,1,2,3,5,6,8,12,20]
            LLAVES_PEDIDO = {0:'restaurante',1:'numero',2:'task_id',3:'repartidor',5:'estado',
                6:'subtotal_pedido',8:'tipo',12:'total_pedido',20:'fecha'}

            try:
                insertar_pedidos(f_pedidos,INDICES_PEDIDO,LLAVES_PEDIDO,dict_pedidos)
                remove_uploaded_file_on(storage_path)
                return redirect('../')
            except Exception as ex:
                return HttpResponse("<h1>Error: Couldn't load data</h1><p>{}</p>".format(ex))
            finally:
                remove_uploaded_file_on(storage_path)
    else:
        form = UploadFileForm()
        return render(request, 'excelempresa/import_intermediate.html', {'form': form, 'modelo':'pedidos'})

def vista_importar_envios(request):
    if request.method  == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            storage_path = 'excelempresa/upload/'
            handle_uploaded_file(request.FILES['file'],storage_path)

            data = openpyxl.load_workbook(storage_path + str(request.FILES['file']), data_only=True)
            t_envios = data["ENVIOS"]

            f_envios = t_envios.max_row - 1
            c_envios = t_envios.max_column

            dict_envios = cells_to_dict(t_envios,f_envios,c_envios)

            INDICES_ENVIO = [0,2,3,6,8]
            LLAVES_ENVIO = {0:'pedido',2:'repartidor',3:'coste_envio',6:'fecha',
                8:'distancia_metros'}

            try:
                insertar_envios(f_envios,INDICES_ENVIO,LLAVES_ENVIO,dict_envios)
                remove_uploaded_file_on(storage_path)
                return redirect('../')
            except Exception as ex:
                return HttpResponse("<h1>Error: Couldn't load data</h1><p>{}</p>".format(ex))
            finally:
                remove_uploaded_file_on(storage_path)
    else:
        form = UploadFileForm()
        return render(request, 'excelempresa/import_intermediate.html', {'form': form, 'modelo':'envios'})

def vista_definir_fecha(request):
    if request.method == "POST":
        form = FechaForm(request.POST)
        if form.is_valid():
            try:
                q = Constants.objects.get(id=1)
            except:
                q = Constants()
                q.save()
            
            q.fecha_inicio = form.cleaned_data['fecha_inicio']
            q.fecha_fin = form.cleaned_data['fecha_fin']
            q.save()
            return redirect('../')

    else:
        try:
            q = Constants.objects.get(id=1)
        except:
            q = Constants()
            q.save()
        
        form = FechaForm()
        return render(request, 'excelempresa/define_fecha.html', {'form': form, 'constants': q})
